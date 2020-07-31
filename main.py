import sqlite3
import csv
import json
from openpyxl import Workbook


def main():
    conn = sqlite3.connect("db.db")
    cursor = conn.cursor()
    class Facility():
        def __init__(self, name):
            self.name = name

        def ust_name(self):
            return self.name

        def vhod_dict(self, vhod_dict):
            self.vhod_dict1 = vhod_dict
            print ("Словарь входных потоков", self.vhod_dict1)

        def vihod_dict(self, vihod_dict):
            self.vihod_dict1 = vihod_dict
            print ("Словарь выходных потоков", self.vihod_dict1)

    class AVT(Facility):
        def display_info(self):
            print("Данные установки: Имя - ", self.name,"Максимальная загрузка - " , self.__load_max,"Словарь входных потоков:", self.vhod_dict1, "Словарь входных потоков:", self.vihod_dict1)

        def load_max(self):
            return self.__load_max

        def load_max1(self, load_max):
            self.__load_max = load_max

    class Secondary(Facility):
        def display_info(self):
            print("Данные установки: Имя - ", self.name,"Максимальная загрузка - " , self.__load_max,"Словарь входных потоков:", self.vhod_dict1, "Словарь входных потоков:", self.vihod_dict1)

        def load_max(self):
            return self.__load_max

        def load_max1(self, load_max):
            self.__load_max = load_max

 #взять установки из БД
    unit_klass_list = []
    unit_max_list = []
    cursor.execute("""SELECT name
                                FROM unit""")
    ustanov_list = cursor.fetchall()
    avt_count = 0
    not_avt_count = 0
    ust_vhod_dict_itog = {}
    ust_vihod_dict_itog = {}

    for ust in ustanov_list:
        srez_ust = str(ust)
        srez_ust = srez_ust[1:len(srez_ust) - 2]  # название установки c кавычками для БД
        srez_ust_for_class = srez_ust[1:len(srez_ust) - 1]  # минус кавычки, норм названия
        cursor.execute("""SELECT value
                               FROM load_max
                               WHERE unit_id = (
                               SELECT id
                               FROM unit
                               WHERE name =""" + srez_ust + """)""")
        ustanov_max = str(cursor.fetchall())
        ustanov_max_int = int(ustanov_max.strip("[(,)]"))
        unit_max_list.append(ustanov_max_int)
        # получение получение типа установки
        cursor.execute("""SELECT type
                               FROM unit
                               WHERE unit.name = """ + srez_ust)
        ustanov_type = cursor.fetchall()
        ustanov_type2 = str(ustanov_type).strip("[(,)]")
   #создание объектов установок
        if ustanov_type2 == str(0):
            avt_count = avt_count + 1
            unit_klass_list.append(AVT(srez_ust_for_class))
        else:
            not_avt_count = not_avt_count + 1
            unit_klass_list.append(Secondary(srez_ust_for_class))
    print("Создано ", avt_count, "установок АВТ")
    print("Создано ", not_avt_count, "установок вторичного производства")
    # наполнение максимального значения с момощью метода
    for alone_ust in unit_klass_list:
        ust_name = "'" + alone_ust.ust_name() + "'"
        cursor.execute("""SELECT value
                                        FROM load_max
                                        WHERE unit_id = (
                                        SELECT id
                                        FROM unit
                                        WHERE name =""" + ust_name + """)""")
        ustanov_max = str(cursor.fetchall())
        ustanov_max_int = int(ustanov_max.strip("[(,)]"))
        alone_ust.load_max1(ustanov_max_int)

#Задача №2
    class Streams:
        def __init__(self, name, start_place, end_place):
            self.name = name 
            self.start_place = start_place
            self.end_place = end_place

        def display_info(self):
            print("Потоки(Имя, Куда входит, Откуда выходит):", self.name, self.start_place, self.end_place)

        def stream_name(self):
            return (self.name)

    #взять потоки из БД
    cursor.execute("""SELECT name 
                            FROM stream""")         # взял снизу!!!!
    stream_list = cursor.fetchall()
    stream_klass_list = []
    count_stream = 0
    for row in stream_list:
        count_stream = count_stream + 1
        srez_stream = str(row)
        srez_stream = srez_stream[1:len(srez_stream)-2]  # без квадратных скобок
        srez_stream_for_class = srez_stream[1:len(srez_stream)-1]  # без кавычек
        #вход
        cursor.execute("""SELECT unit.name 
                                    FROM unit_material
                                    LEFT JOIN unit ON unit_material.unit_id = unit.id
                                    LEFT JOIN stream ON unit_material.stream_id = stream.id
                                    WHERE unit_material.feed_flag = 1 
                                    AND stream.name = """ + srez_stream)       #установки в которые входит
        streams_kazdiy_vhod = cursor.fetchall()
        streams_kazdiy_vhod1 = []
        for every in streams_kazdiy_vhod:
            every1 = str(every).strip("('',)")
            streams_kazdiy_vhod1.append(every1)
        #выход
        cursor.execute("""SELECT unit.name 
                                            FROM unit_material
                                            LEFT JOIN unit ON unit_material.unit_id = unit.id
                                            LEFT JOIN stream ON unit_material.stream_id = stream.id
                                            WHERE unit_material.feed_flag = 0 
                                            AND stream.name = """ + srez_stream)  # установки с которых выходит
        streams_kazdiy_vihod = cursor.fetchall()
        streams_kazdiy_vihod1 = []
        for every2 in streams_kazdiy_vihod:
            every21 = str(every2).strip("('',)")
            streams_kazdiy_vihod1.append(every21)
        stream_klass_list.append(Streams(srez_stream_for_class, streams_kazdiy_vhod1, streams_kazdiy_vihod1)) # добавление потока в список потоков
    # когда потоки созданы, добавляем их в словари
    for ust in ustanov_list:  # для каждой установки
        srez_ust = str(ust)
        srez_ust = srez_ust[1:len(srez_ust) - 2]  # название установки c кавычками для БД
        srez_ust_for_class = srez_ust[1:len(srez_ust) - 1]  # минус кавычки, норм названия
        cursor.execute("""SELECT stream.name 
                                       FROM unit_material
                                       LEFT JOIN unit ON unit_material.unit_id = unit.id
                                       LEFT JOIN stream ON unit_material.stream_id = stream.id
                                       WHERE unit_material.feed_flag = 1 
                                       AND unit.name = """ + srez_ust)
        ust_vhod = cursor.fetchall()
        for ust_vhod2 in ust_vhod:      # для каждого входа
            ust_vhod_dict = str(ust_vhod2).strip("(',')")
            for alone_stream in stream_klass_list:
                if alone_stream.stream_name() == ust_vhod_dict:
                    ust_vhod_dict_itog[ust_vhod_dict] = alone_stream
        print("Установка:", srez_ust_for_class)
        print("Словарь входных потоков:", ust_vhod_dict_itog)

        # а теперь выходные
        cursor.execute("""SELECT stream.name 
                                               FROM unit_material
                                               LEFT JOIN unit ON unit_material.unit_id = unit.id
                                               LEFT JOIN stream ON unit_material.stream_id = stream.id
                                               WHERE unit_material.feed_flag = 0 
                                               AND unit.name = """ + srez_ust)
        ust_vihod = cursor.fetchall()
        for ust_vihod2 in ust_vihod:  # для каждого входа
            ust_vihod_dict = str(ust_vihod2).strip("(',')")
            for alone_stream in stream_klass_list:
                if alone_stream.stream_name() == ust_vihod_dict:
                    ust_vihod_dict_itog[ust_vihod_dict] = alone_stream
        print("Словарь выходных потоков:", ust_vihod_dict_itog)
        # добавление словаря входных потоков в класс
        for alone_ust in unit_klass_list:
            if alone_ust.ust_name() == srez_ust_for_class:
                alone_ust.vhod_dict(ust_vhod_dict_itog)
                alone_ust.vihod_dict(ust_vihod_dict_itog)

    for alone_ust in unit_klass_list: #выдать что лежит в объекте установки
        alone_ust.display_info()

    for alone_stream in stream_klass_list:  #выдать что лежит в объекте потока
        alone_stream.display_info()

#Задача №3:
    cursor.execute("""SELECT unit.name, stream.name 
                    FROM unit_material
					LEFT JOIN unit ON unit_material.unit_id = unit.id
					LEFT JOIN stream ON unit_material.stream_id = stream.id
                    WHERE unit_material.feed_flag = 1""")
    unit_list = cursor.fetchall()
    print("Задача №3 Ответ:", unit_list)

#Задача №4
    cursor.execute("""SELECT *
                        FROM stream
                        WHERE stream.id 
                        NOT IN 
						(SELECT stream_id
						FROM unit_material) """)
    unit_list2 = cursor.fetchall()
    print("Задача №4 Ответ:", unit_list2)

    csvfile = "data4.csv"
    with open(csvfile, "w", newline="") as file:
        writer = csv.writer(file)
        writer.writerows(unit_list2)

# Задача №5
    resultdict = {}
    for row in stream_list:
        srez_stream = str(row)
        srez_stream = srez_stream[1:len(srez_stream)-2] # я сдаюсь, не знаю как получить из БД строку нормально
        cursor.execute("""SELECT unit.name 
                            FROM unit_material
                            LEFT JOIN unit ON unit_material.unit_id = unit.id
                            LEFT JOIN stream ON unit_material.stream_id = stream.id
                            WHERE unit_material.feed_flag = 1 
                            AND stream.name = """+ srez_stream)
        streams_kazdiy = cursor.fetchall()
        if len(streams_kazdiy) > 1:
            resultdict[srez_stream] = streams_kazdiy
    print("Задача №5 Ответ:", resultdict)

    with open('data5.json', 'w', encoding='utf-8') as fh:  # кусок про json
        fh.write(
            json.dumps(resultdict, ensure_ascii=False))  # словарь в unicode-строку

    # Задача №6
    wb = Workbook()
    cursor.execute("""SELECT name
                            FROM unit""")
    ustanov_list = cursor.fetchall()
    print("Задача №6 Ответ:")
    for ust in ustanov_list:
        srez_ust = str(ust)
        srez_ust = srez_ust[1:len(srez_ust) - 2]  # название установки
        srez_ust_excel = srez_ust[1:len(srez_ust) - 1]  # минус кавычки
        cursor.execute("""SELECT stream.name 
                                    FROM unit_material
                                    LEFT JOIN unit ON unit_material.unit_id = unit.id
                                    LEFT JOIN stream ON unit_material.stream_id = stream.id
                                    WHERE unit_material.feed_flag = 1 
                                    AND unit.name = """ + srez_ust)
        ust_vhod = cursor.fetchall()
        print("Вход", srez_ust_excel, ust_vhod)
        cursor.execute("""SELECT stream.name 
                                            FROM unit_material
                                            LEFT JOIN unit ON unit_material.unit_id = unit.id
                                            LEFT JOIN stream ON unit_material.stream_id = stream.id
                                            WHERE unit_material.feed_flag = 0 
                                            AND unit.name = """ + srez_ust)
        ust_vihod = cursor.fetchall()
        print("Выход", srez_ust_excel, ust_vihod)
        ws1 = wb.create_sheet(srez_ust_excel)
        ws1['A1']="Вход"
        ws1['B1'] = "Выход"
        for x in range(1, len(ust_vhod)+1): # x от 1 до длины кол-ва потоков+1 так как сдвигается на 1 т.к. эксель начинает с 1 а не с 0, +1 т.к.
            znach1 = str(ust_vhod[x-1])     #в 1 ставим значение под 0 и тд
            znach1 = znach1[2:len(znach1) - 3]  #режем чтоб красиво вставить
            ws1.cell(row=x+1, column=1, value=znach1)
        for x in range(1, len(ust_vihod)+1):
            znach2 = str(ust_vihod[x-1])
            znach2 = znach2[2:len(znach2) - 3]
            ws1.cell(row=x+1, column=2, value=znach2)
    wb.remove(wb['Sheet'])
    wb.save('data6.xlsx')
    conn.close()

if __name__ == "__main__":
    main()